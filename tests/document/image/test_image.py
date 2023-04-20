# coding: utf-8
from io import BytesIO

import pytest
from openpyxl import load_workbook

from img2table.document.image import Image
from img2table.ocr import TesseractOCR
from img2table.tables.objects.extraction import BBox


def test_validators():
    with pytest.raises(TypeError) as e_info:
        img = Image(src=1)

    with pytest.raises(TypeError) as e_info:
        img = Image(src="img", detect_rotation=3)


def test_load_image():
    # Load from path
    img_from_path = Image(src="test_data/test.png")

    # Load from bytes
    with open("test_data/test.png", "rb") as f:
        img_from_bytes = Image(src=f.read())

    # Load from BytesIO
    with open("test_data/test.png", "rb") as f:
        img_from_bytesio = Image(src=BytesIO(f.read()))

    assert img_from_path.bytes == img_from_bytes.bytes == img_from_bytesio.bytes

    assert list(img_from_path.images)[0].shape == (417, 1365)


def test_blank_image(mock_tesseract):
    ocr = TesseractOCR()
    img = Image(src="test_data/blank.png",
                detect_rotation=True)

    result = img.extract_tables(ocr=ocr,
                                implicit_rows=True,
                                borderless_tables=True,
                                min_confidence=50)

    assert result == []


def test_blank_no_ocr():
    img = Image(src="test_data/blank.png",
                detect_rotation=True)

    result = img.extract_tables(implicit_rows=True,
                                borderless_tables=True,
                                min_confidence=50)

    assert result == []


def test_image_tables(mock_tesseract):
    ocr = TesseractOCR()
    img = Image(src="test_data/test.png",
                detect_rotation=True)

    result = img.extract_tables(ocr=ocr, implicit_rows=True, min_confidence=50)

    assert len(result) == 2

    assert result[0].title is None
    assert result[0].bbox == BBox(x1=35, y1=20, x2=770, y2=327)
    assert len(result[0].content) == 6
    assert len(result[0].content[0]) == 3

    assert result[1].title is None
    assert result[1].bbox == BBox(x1=962, y1=21, x2=1154, y2=123)
    assert len(result[1].content) == 2
    assert len(result[1].content[0]) == 2


def test_no_ocr():
    img = Image(src="test_data/dark.png",
                detect_rotation=True)

    result = img.extract_tables(implicit_rows=True, min_confidence=50)

    assert len(result) == 1

    assert result[0].title is None
    assert result[0].bbox == BBox(x1=36, y1=40, x2=840, y2=529)
    assert len(result[0].content) == 19
    assert len(result[0].content[0]) == 7


def test_image_excel(mock_tesseract):
    ocr = TesseractOCR()
    img = Image(src="test_data/test.png",
                detect_rotation=True)

    result = img.to_xlsx(dest=BytesIO(), ocr=ocr, implicit_rows=True, min_confidence=50)

    expected = load_workbook(filename="test_data/expected.xlsx")
    result_wb = load_workbook(filename=result)

    for idx, ws in enumerate(result_wb.worksheets):
        assert ws.title == expected.worksheets[idx].title
        assert list(ws.values) == list(expected.worksheets[idx].values)
